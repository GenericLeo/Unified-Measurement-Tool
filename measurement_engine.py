"""
Measurement Engine Module
Handles measurement calculations, statistics, and data export
"""

import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple
import numpy as np


def calculate_segment_statistics(segments: Dict[int, List[Tuple[int, int]]]) -> dict:
    """
    Calculate statistics for all segments.
    
    Args:
        segments: Dictionary mapping coordinates to list of segments
        
    Returns:
        Dictionary containing statistics (count, min, max, mean, median)
    """
    all_lengths = []
    
    for coord, seg_list in segments.items():
        for start, end in seg_list:
            length = end - start + 1
            all_lengths.append(length)
    
    if not all_lengths:
        return {
            'count': 0,
            'min': 0,
            'max': 0,
            'mean': 0,
            'median': 0,
            'std_dev': 0
        }
    
    return {
        'count': len(all_lengths),
        'min': min(all_lengths),
        'max': max(all_lengths),
        'mean': np.mean(all_lengths),
        'median': np.median(all_lengths),
        'std_dev': np.std(all_lengths)
    }


def save_vertical_segments_to_csv(images: List[str], processed_images: List, 
                                   csv_path: str) -> None:
    """
    Save vertical segment data to CSV file.
    
    Args:
        images: List of image names
        processed_images: List of processed binary images
        csv_path: Path where to save the CSV file
    """
    from image_processor import analyze_all_vertical_segments
    
    with open(csv_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        # Write header
        writer.writerow(['Image Name', 'Column (X)', 'Segment Start (Y)', 
                        'Segment End (Y)', 'Segment Length (pixels)'])
        
        # Process each image
        for img_name, bw in zip(images, processed_images):
            if bw is not None:
                # Get all segments for this image
                all_segments = analyze_all_vertical_segments(bw)
                
                # Write each segment to CSV
                for x, segments in sorted(all_segments.items()):
                    for start_y, end_y in segments:
                        length = end_y - start_y + 1
                        writer.writerow([img_name, x, start_y, end_y, length])


def save_horizontal_segments_to_csv(images: List[str], processed_images: List, 
                                     csv_path: str) -> None:
    """
    Save horizontal segment data to CSV file.
    
    Args:
        images: List of image names
        processed_images: List of processed binary images
        csv_path: Path where to save the CSV file
    """
    from image_processor import analyze_all_horizontal_segments
    
    with open(csv_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        # Write header
        writer.writerow(['Image Name', 'Row (Y)', 'Segment Start (X)', 
                        'Segment End (X)', 'Segment Length (pixels)'])
        
        # Process each image
        for img_name, bw in zip(images, processed_images):
            if bw is not None:
                # Get all segments for this image
                all_segments = analyze_all_horizontal_segments(bw)
                
                # Write each segment to CSV
                for y, segments in sorted(all_segments.items()):
                    for start_x, end_x in segments:
                        length = end_x - start_x + 1
                        writer.writerow([img_name, y, start_x, end_x, length])


def save_vertical_segments_to_excel(images: List[str], processed_images: List, 
                                     excel_path: str) -> None:
    """
    Save vertical segment data to Excel file with separate sheets per image.
    
    Args:
        images: List of image names
        processed_images: List of processed binary images
        excel_path: Path where to save the Excel file
    """
    from image_processor import analyze_all_vertical_segments
    
    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Process each image
    for img_name, bw in zip(images, processed_images):
        if bw is not None:
            # Create a safe sheet name (max 31 chars, no special chars)
            sheet_name = img_name[:31].replace('/', '_').replace('\\', '_').replace(':', '_')
            ws = wb.create_sheet(title=sheet_name)
            
            # Write header
            ws.append(['Column (X)', 'Segment Start (Y)', 'Segment End (Y)', 
                      'Segment Length (pixels)'])
            
            # Get all segments for this image
            all_segments = analyze_all_vertical_segments(bw)
            
            # Write each segment
            for x, segments in sorted(all_segments.items()):
                for start_y, end_y in segments:
                    length = end_y - start_y + 1
                    ws.append([x, start_y, end_y, length])
            
            # Auto-adjust column widths
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save the workbook
    wb.save(excel_path)


def save_horizontal_segments_to_excel(images: List[str], processed_images: List, 
                                       excel_path: str) -> None:
    """
    Save horizontal segment data to Excel file with separate sheets per image.
    
    Args:
        images: List of image names
        processed_images: List of processed binary images
        excel_path: Path where to save the Excel file
    """
    from image_processor import analyze_all_horizontal_segments
    
    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Process each image
    for img_name, bw in zip(images, processed_images):
        if bw is not None:
            # Create a safe sheet name (max 31 chars, no special chars)
            sheet_name = img_name[:31].replace('/', '_').replace('\\', '_').replace(':', '_')
            ws = wb.create_sheet(title=sheet_name)
            
            # Write header
            ws.append(['Row (Y)', 'Segment Start (X)', 'Segment End (X)', 
                      'Segment Length (pixels)'])
            
            # Get all segments for this image
            all_segments = analyze_all_horizontal_segments(bw)
            
            # Write each segment
            for y, segments in sorted(all_segments.items()):
                for start_x, end_x in segments:
                    length = end_x - start_x + 1
                    ws.append([y, start_x, end_x, length])
            
            # Auto-adjust column widths
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save the workbook
    wb.save(excel_path)


def pixels_to_physical_measurement(pixel_length: float, scale_factor: float, 
                                    unit: str = "μm") -> Tuple[float, str]:
    """
    Convert pixel measurements to physical measurements.
    
    Args:
        pixel_length: Length in pixels
        scale_factor: Conversion factor (physical units per pixel)
        unit: Physical unit (default "μm")
    
    Returns:
        Tuple of (physical_length, unit_string)
    """
    physical_length = pixel_length * scale_factor
    return physical_length, unit


def write_interface_distances_csv(out_path: str, rows: List[Dict[str, object]]) -> None:
    """Write interface distance measurements to CSV file."""
    if not rows:
        with open(out_path, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["mask_file", "X Position (px)", "Interface Y (px)", "Distance (px)"])
        return

    all_keys = set().union(*[r.keys() for r in rows])
    preferred = [
        "mask_file", "PixelSize (nm/px)",
        "X Position (px)", "Interface Y (px)", "Distance to Top (px)", "Distance to Bottom (px)",
        "X Position (um)", "Interface Y (um)", "Distance to Top (um)", "Distance to Bottom (um)",
    ]
    header = [k for k in preferred if k in all_keys] + [k for k in sorted(all_keys) if k not in set(preferred)]

    with open(out_path, "w", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_measurements_csv(out_path: str, rows: List[Dict[str, object]]) -> None:
    """Write carbide feature measurements to CSV file."""
    if not rows:
        with open(out_path, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["mask_file", "PixelSize (nm/px)"])
        return

    preferred = [
        "mask_file", "PixelSize (nm/px)",
        "Area (px^2)", "Aspect Ratio", "Centroid X (px)", "Centroid Y (px)",
        "Equivalent Diameter (px)", "Extent", "Height (px)", "Major Axis (px)",
        "Minor Axis (px)", "Orientation (deg)", "Perimeter (px)", "width (px)", "X (px)", "Y (px)",
        "Baseline Y at Centroid (px)", "Normalized Centroid Y (px)", "Normalized Y (px)",
        "Area (um^2)", "Perimeter (um)", "Equivalent Diameter (um)", "width (um)", "Height (um)",
        "Centroid X (um)", "Centroid Y (um)", "Major Axis (um)", "Minor Axis (um)",
        "Baseline Y at Centroid (um)", "Normalized Centroid Y (um)", "Normalized Y (um)",
    ]
    all_keys = set().union(*[r.keys() for r in rows])
    header = [k for k in preferred if k in all_keys] + [k for k in sorted(all_keys) if k not in set(preferred)]

    with open(out_path, "w", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_interface_distances_excel(out_path: str, rows: List[Dict[str, object]]) -> None:
    """Write interface distance measurements to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Interface Distances"

    if not rows:
        ws.append(["mask_file", "X Position (px)", "Interface Y (px)", "Distance (px)"])
        wb.save(out_path)
        return

    all_keys = set().union(*[r.keys() for r in rows])
    preferred = [
        "mask_file", "PixelSize (nm/px)",
        "X Position (px)", "Interface Y (px)", "Distance to Top (px)", "Distance to Bottom (px)",
        "X Position (um)", "Interface Y (um)", "Distance to Top (um)", "Distance to Bottom (um)",
    ]
    header = [k for k in preferred if k in all_keys] + [k for k in sorted(all_keys) if k not in set(preferred)]
    ws.append(header)

    for row_dict in rows:
        ws.append([row_dict.get(k, "") for k in header])

    wb.save(out_path)


def write_measurements_excel(out_path: str, rows: List[Dict[str, object]]) -> None:
    """Write carbide feature measurements to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Measurements"

    if not rows:
        ws.append(["mask_file", "PixelSize (nm/px)"])
        wb.save(out_path)
        return

    preferred = [
        "mask_file", "PixelSize (nm/px)",
        "Area (px^2)", "Aspect Ratio", "Centroid X (px)", "Centroid Y (px)",
        "Equivalent Diameter (px)", "Extent", "Height (px)", "Major Axis (px)",
        "Minor Axis (px)", "Orientation (deg)", "Perimeter (px)", "width (px)", "X (px)", "Y (px)",
        "Baseline Y at Centroid (px)", "Normalized Centroid Y (px)", "Normalized Y (px)",
        "Area (um^2)", "Perimeter (um)", "Equivalent Diameter (um)", "width (um)", "Height (um)",
        "Centroid X (um)", "Centroid Y (um)", "Major Axis (um)", "Minor Axis (um)",
        "Baseline Y at Centroid (um)", "Normalized Centroid Y (um)", "Normalized Y (um)",
    ]
    all_keys = set().union(*[r.keys() for r in rows])
    header = [k for k in preferred if k in all_keys] + [k for k in sorted(all_keys) if k not in set(preferred)]
    ws.append(header)

    for row_dict in rows:
        ws.append([row_dict.get(k, "") for k in header])

    wb.save(out_path)
